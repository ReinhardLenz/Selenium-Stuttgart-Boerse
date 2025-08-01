import os
import time
import csv
import random
import datetime
import warnings
import numpy as np
import pandas as pd
import clipboard as clp
import openpyxl
import winsound

from bs4 import BeautifulSoup
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager

import config_webdriver_manager
import web_actions
from locators import Locators

# Allow file download
config_webdriver_manager.driver.command_executor._commands["send_command"] = (
    "POST", "/session/$sessionId/chromium/send_command"
)
params = {
    'cmd': 'Page.setDownloadBehavior',
    'params': {'behavior': 'allow', 'downloadPath': "/path/to/download/dir"}
}
config_webdriver_manager.driver.execute("send_command", params)

# Open main page
config_webdriver_manager.driver.get("https://www.boerse-stuttgart.de/de-de/tools/produktsuche/anleihen-finder/")
time.sleep(random.randint(2, 3))

# Accept cookies
button = config_webdriver_manager.driver.find_element(By.CSS_SELECTOR, Locators.cookie_acceptor_css_new)
button.click()

web_actions.zoom40androlldown()
web_actions.relax()
config_webdriver_manager.driver.execute_script("document.body.style.zoom='60%'")
web_actions.relax()

# Select currency filter
web_actions.click_operation_xpath(Locators.Wahrung_click_XPATH)
web_actions.click_operation_xpath(Locators.Wahrung_text_field_click_XPATH)
web_actions.write_operation_selector(Locators.Wahrung_text_field_click_CSS_3_SELECTOR, "Euro")
time.sleep(random.randint(2, 6))

checkbox = config_webdriver_manager.driver.find_element(By.CSS_SELECTOR, Locators.Wahrung_checkbox_Eur1_CSS)
checkbox.click()
web_actions.click_operation_xpath(Locators.Wahrung_Anwenden_XP)
print("Milestone Click Anwenden of euro through")

time.sleep(random.randint(2, 6))
web_actions.relax()

# Apply additional filters
web_actions.click_operation_Xpath(Locators.Zusatzliche_Filter_right_XPATH)
web_actions.click_operation_css(Locators.handelbare_Einh_css)
web_actions.click_operation_css(Locators.Gelisteter_Zeitraum_Button_css)
web_actions.relax()
time.sleep(random.randint(2, 6))

web_actions.click_operation_Xpath(Locators.Zusatzliche_Filter_Anwenden_Full_XPath)
web_actions.relax()
web_actions.click_operation_Xpath(Locators.Handelbare_einh_Button_xpath)
web_actions.relax()
time.sleep(random.randint(2, 6))

# Min/Max Littera
web_actions.clear_operation_A_xpath(Locators.min_Littera_xp)
web_actions.write_operation_b_xpath(Locators.min_Littera_xp, "1")
time.sleep(random.randint(2, 6))
web_actions.clear_operation_A_xpath(Locators.max_Littera_xp)

input_field = config_webdriver_manager.driver.find_element(By.CSS_SELECTOR, Locators.max_Littera_css)
input_field.send_keys(Keys.CONTROL + "a")
input_field.send_keys(Keys.DELETE)
input_field.send_keys('1000', Keys.RETURN)

web_actions.click_operation_Xpath(Locators.Anwenden_littera_xp)

# Maturity range
Smaller_date = datetime.date.today() + relativedelta(years=5)
Smaller_date_format = Smaller_date.strftime("%d.%m.%Y")
web_actions.write_to_input_field_css_auslaufzeit_AI(Smaller_date_format, Locators.Auslaufzeit_AI_css, 0)

Bigger_date = datetime.date.today() + relativedelta(years=10)
Bigger_date_format = Bigger_date.strftime("%d.%m.%Y")
web_actions.write_to_input_field_css_auslaufzeit_AI(Bigger_date_format, Locators.Auslaufzeit_AI_css, 1)

# Listed date
date_listed = datetime.date.today() - relativedelta(years=3)
date_format_listed = date_listed.strftime("%d.%m.%Y")

wait = WebDriverWait(config_webdriver_manager.driver, 10)
input_field_date_listed = wait.until(
    EC.element_to_be_clickable((By.XPATH, Locators.Gelisteter_Zeitraum_datum_field_FULL_XPATH))
)
input_field_date_listed.click()
time.sleep(random.randint(1, 3))
input_field_date_listed.send_keys(date_format_listed, Keys.RETURN)
time.sleep(random.randint(1, 3))

# Bond type
web_actions.click_operation_Xpath(Locators.Preselector_Anleihen_Typ_FULL_XPATH)
web_actions.relax()
web_actions.click_operation_Xpath(Locators.Unternehmensanleihe_type_XPATH)
web_actions.relax()
web_actions.write_operation_xpath(Locators.Unternehmensanleihe_type_XPATH, "Unternehmensanleihe")
time.sleep(random.randint(2, 4))
web_actions.relax()

xpath_expression = Locators.Unternehmensanleihe_type_XPATH_2
element = WebDriverWait(config_webdriver_manager.driver, 20).until(
    EC.presence_of_element_located((By.XPATH, xpath_expression))
)
actions = ActionChains(config_webdriver_manager.driver)
time.sleep(random.randint(1, 2))
web_actions.relax()
actions.move_to_element(element).click().perform()

web_actions.click_operation_Xpath(Locators.Anwenden_Unternehmensanleihe_XPATH)
web_actions.relax()
time.sleep(random.randint(1, 2))
web_actions.click_operation_Xpath(Locators.Rendite_button_XPATH)

time.sleep(random.randint(1, 2))
inputs = WebDriverWait(config_webdriver_manager.driver, 10).until(
    EC.presence_of_all_elements_located((By.CSS_SELECTOR, Locators.Rendite_text_field_CSS_SELECTOR))
)
inputs[0].click()
inputs[0].send_keys(Keys.CONTROL + "a")
inputs[0].send_keys(Keys.DELETE)
inputs[0].send_keys("3%")
time.sleep(random.randint(1, 2))

anwenden_button = WebDriverWait(config_webdriver_manager.driver, 10).until(
    EC.element_to_be_clickable((By.CSS_SELECTOR, Locators.Rendite_Anwenden_CSS_SELECTOR))
)
anwenden_button.click()

# Scroll to bottom
config_webdriver_manager.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")

# Parse result table
html_source = config_webdriver_manager.driver.page_source
soup = BeautifulSoup(html_source, 'html.parser')
table = soup.table
table_rows = table.find_all('tr')

tables = []
for tr in table_rows:
    td = tr.find_all('td')
    row = []
    for i in td:
        word = i.text.strip().replace(".000", "000").replace(",000", "")
        row.append(word)
    tables.append(row)

print(tables)
df = pd.DataFrame(tables)

# Load and update Excel
wb = load_workbook('c:\\temp\\tutorial\\test1.xlsx')
print(wb.sheetnames)
print("Workbook loaded successfully")

sheet = wb.active
if sheet is None:
    raise ValueError("Worksheet is not loaded correctly.")

# Insert 10 columns at beginning
for _ in range(10):
    sheet.insert_cols(1)

# Write dataframe to Excel and set headers
web_actions.df_to_excel(df, wb.active)
today = datetime.date.today()
sheet["A2"] = today.strftime("%d.%m.%Y")
sheet["A3"] = "Nr"
sheet["B3"] = "ISIN WKN"
sheet["C3"] = "Issuer"
sheet["D3"] = "Coupon"
sheet["E3"] = "Maturity Date"
sheet["F3"] = "Yield"
sheet["G3"] = "Currency"
sheet["H3"] = "not used"
sheet["I3"] = "min littera"
sheet["J3"] = "not used"

wb.save('c:\\temp\\tutorial\\test1.xlsx')
print("Data written to Excel file successfully")

# Close browser
config_webdriver_manager.driver.quit()
